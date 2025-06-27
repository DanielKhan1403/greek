from django.contrib import admin
from django.http import HttpResponse
from .models import BeMemberForm
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.urls import path
from django.core.exceptions import PermissionDenied


@admin.register(BeMemberForm)
class BeMemberFormAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'age', 'email_address', 'phone_number', 'message', 'status')
    list_filter = ('age','status')
    search_fields = ('full_name', 'email_address', 'phone_number')
    actions = ['export_selected_to_excel']
    change_list_template = "change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export/",
                self.admin_site.admin_view(self.export_all_to_excel),
                name="social_works_bememberform_export",
            ),
        ]
        return custom_urls + urls

    # ✅ Action для экспорта выделенных
    def export_selected_to_excel(self, request, queryset):
        return self._generate_excel(queryset)

    export_selected_to_excel.short_description = "Экспортировать выбранные заявки в Excel"

    # ✅ Отдельная view для экспорта всех
    def export_all_to_excel(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied("У тебя нет доступа к экспорту данных.")
        queryset = BeMemberForm.objects.all()
        return self._generate_excel(queryset)

    # ✅ Общая логика генерации Excel
    def _generate_excel(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные заявок"

        headers = ['Полное имя', 'Возраст', 'Эл. почта', 'Номер телефона', 'Сообщение']
        ws.append(headers)

        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for form in queryset:
            ws.append([
                form.full_name,
                form.age,
                form.email_address,
                str(form.phone_number),
                form.message
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Заявки_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)
        return response
