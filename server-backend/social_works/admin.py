from django.contrib import admin
from django.http import HttpResponse
from .models import BeMemberForm
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.urls import path
from django.core.exceptions import PermissionDenied

from django.contrib import messages
@admin.register(BeMemberForm)
class BeMemberFormAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'birth_date','get_age_verbose', 'email', 'phone_number', 'telegram_account','status')
    list_filter = ('status',)
    search_fields = ('full_name', 'email', 'phone_number')
    actions = ['export_selected_to_excel']
    change_list_template = "change_list.html"

    def get_age_verbose(self, obj):
        return obj.age_verbose

    get_age_verbose.short_description = "Возраст"

    def save_model(self, request, obj, form, change):
        if change:
            old_obj = BeMemberForm.objects.get(pk=obj.pk)
            if old_obj.status != obj.status and not request.user.is_superuser:
                messages.error(request, "❌ Только суперпользователь может менять статус.")
                # откатим изменение статуса
                obj.status = old_obj.status
        super().save_model(request, obj, form, change)


    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export/",
                self.admin_site.admin_view(self.export_all_to_excel),
                name="social_works_bememberform_export",
            ),
        ]
        return custom_urls + urls

    def export_selected_to_excel(self, request, queryset):
        return self._generate_excel(queryset)

    export_selected_to_excel.short_description = "Экспортировать выбранные заявки в Excel"

    def export_all_to_excel(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied("У тебя нет доступа к экспорту данных.")
        queryset = BeMemberForm.objects.all()
        return self._generate_excel(queryset)

    def _generate_excel(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные заявок"

        headers = [
            'ФИО', 'Дата рождения', 'Телефон', 'Telegram', 'Email',
            'Адрес', 'Греческий предок (кто)', 'Фамилия предка',
            'О себе', 'Интересы', 'Вклад', 'Статус'
        ]
        ws.append(headers)

        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for form in queryset:
            ws.append([
                form.full_name,
                form.birth_date.strftime('%d.%m.%Y') if form.birth_date else '',
                str(form.phone_number),
                form.telegram_account or '',
                form.email,
                form.address,
                form.greek_ancestor_relation,
                form.greek_ancestor_surname,
                form.about_yourself or '',
                form.interested_activities or '',
                form.your_contribution or '',
                form.get_status_display()
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Заявки_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)
        return response
